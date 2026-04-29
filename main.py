#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gestão de Atividades - Sistema de Gerenciamento de Tarefas para TI
Versão com PyQt6 e System Tray nativo
"""

import sys
import os
import json
import re
import time
import threading
from datetime import datetime, timedelta
from collections import defaultdict

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QLabel, QLineEdit, QPushButton, QComboBox, QTableWidget, 
    QTableWidgetItem, QHeaderView, QMessageBox, QSystemTrayIcon,
    QMenu, QDialog, QGroupBox, QDateEdit, QFileDialog, QAction
)
from PyQt6.QtCore import Qt, QTimer, QDate, pyqtSignal
from PyQt6.QtGui import QIcon, QColor, QPixmap

try:
    from plyer import notification
    HAS_PLYER = True
except ImportError:
    HAS_PLYER = False


class Database:
    """Gerenciador de banco de dados SQLite e JSON"""
    
    def __init__(self, db_file="gestao_atividades.db"):
        self.db_file = db_file
        self.json_file = "tasks_db.json"
        self.tasks = {}
        self.load_data()
    
    def load_data(self):
        """Carrega dados do arquivo JSON"""
        if os.path.exists(self.json_file):
            try:
                with open(self.json_file, "r", encoding="utf-8") as f:
                    self.tasks = json.load(f)
            except Exception as e:
                print(f"Erro ao carregar dados: {e}")
                self.tasks = {}
        else:
            self.tasks = {}
    
    def save_data(self):
        """Salva dados no arquivo JSON"""
        try:
            with open(self.json_file, "w", encoding="utf-8") as f:
                json.dump(self.tasks, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Erro ao salvar dados: {e}")
    
    def add_task(self, task):
        """Adiciona uma tarefa"""
        self.tasks[task['id']] = task
        self.save_data()
    
    def update_task(self, task_id, task_data):
        """Atualiza uma tarefa"""
        if task_id in self.tasks:
            self.tasks[task_id].update(task_data)
            self.save_data()
    
    def delete_task(self, task_id):
        """Exclui uma tarefa"""
        if task_id in self.tasks:
            del self.tasks[task_id]
            self.save_data()
    
    def get_task(self, task_id):
        """Obtém uma tarefa pelo ID"""
        return self.tasks.get(task_id)
    
    def get_all_tasks(self):
        """Obtém todas as tarefas"""
        return self.tasks
    
    def get_completed_tasks(self):
        """Obtém apenas tarefas concluídas"""
        return {k: v for k, v in self.tasks.items() if v.get('status') == 'Concluido'}


class TaskParser:
    """Parser para linha de chamado"""
    
    @staticmethod
    def parse(line):
        """
        Parse da linha do chamado
        Formato: #TI2604276854\temail\tcategoria\t"descrição"
        """
        line = line.strip()
        if not line:
            return None
        
        # Tenta separar por tabulação
        parts = line.split('\t')
        
        if len(parts) >= 4:
            ticket = parts[0].strip()
            email = parts[1].strip()
            category = parts[2].strip()
            description = parts[3].strip().strip('"')
            
            return {
                'ticket': ticket,
                'email': email,
                'category': category,
                'description': description
            }
        
        # Fallback: regex para formato sem tabs
        match = re.match(r'(#\w+)\s+(\S+)\s+(.*?)\s+"(.*)"', line)
        if match:
            return {
                'ticket': match.group(1),
                'email': match.group(2),
                'category': match.group(3),
                'description': match.group(4)
            }
        
        return None


class LogsDialog(QDialog):
    """Janela de logs com filtragem"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Histórico de Logs - Gestão de Atividades")
        self.setMinimumSize(900, 600)
        self.setup_ui()
        self.load_logs()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Frame de filtros
        filter_frame = QGroupBox("Filtros")
        filter_layout = QHBoxLayout(filter_frame)
        
        # Filtro de texto
        filter_layout.addWidget(QLabel("Texto:"))
        self.text_filter = QLineEdit()
        self.text_filter.setPlaceholderText("Buscar por ticket, email, descrição...")
        filter_layout.addWidget(self.text_filter)
        
        # Filtro de data inicial
        filter_layout.addWidget(QLabel("De:"))
        self.date_start = QDateEdit()
        self.date_start.setCalendarPopup(True)
        self.date_start.setDate(QDate.currentDate().addMonths(-1))
        filter_layout.addWidget(self.date_start)
        
        # Filtro de data final
        filter_layout.addWidget(QLabel("Até:"))
        self.date_end = QDateEdit()
        self.date_end.setCalendarPopup(True)
        self.date_end.setDate(QDate.currentDate())
        filter_layout.addWidget(self.date_end)
        
        # Botão filtrar
        btn_filter = QPushButton("🔍 Filtrar")
        btn_filter.clicked.connect(self.load_logs)
        filter_layout.addWidget(btn_filter)
        
        # Botão exportar
        btn_export = QPushButton("📥 Exportar Excel")
        btn_export.clicked.connect(self.export_excel)
        filter_layout.addWidget(btn_export)
        
        layout.addWidget(filter_frame)
        
        # Tabela de logs
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "Ticket", "Email", "Categoria", "Prioridade", 
            "Data Conclusão", "Tempo Total", "Descrição"
        ])
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(0, 120)  # Ticket
        self.table.setColumnWidth(1, 200)  # Email
        self.table.setColumnWidth(2, 150)  # Categoria
        self.table.setColumnWidth(3, 80)   # Prioridade
        self.table.setColumnWidth(4, 120)  # Data
        self.table.setColumnWidth(5, 100)  # Tempo
        self.table.setColumnWidth(6, 300)  # Descrição
        
        layout.addWidget(self.table)
        
        # Botão fechar
        btn_close = QPushButton("Fechar")
        btn_close.clicked.connect(self.close)
        layout.addWidget(btn_close)
    
    def load_logs(self):
        """Carrega logs na tabela com filtros"""
        self.table.setRowCount(0)
        
        text = self.text_filter.text().lower()
        date_start = self.date_start.date().toPyDate()
        date_end = self.date_end.date().toPyDate()
        
        tasks = self.db.get_completed_tasks()
        
        for task_id, task in tasks.items():
            # Aplicar filtro de texto
            if text:
                searchable = f"{task.get('ticket', '')} {task.get('email', '')} {task.get('description', '')}".lower()
                if text not in searchable:
                    continue
            
            # Aplicar filtro de data
            end_time_str = task.get('end_time', '')
            if end_time_str:
                try:
                    end_date = datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S").date()
                    if end_date < date_start or end_date > date_end:
                        continue
                except:
                    pass
            
            # Adicionar linha
            row = self.table.rowCount()
            self.table.insertRow(row)
            
            # Formatar tempo
            total_seconds = task.get('total_seconds', 0)
            hours = int(total_seconds // 3600)
            minutes = int((total_seconds % 3600) // 60)
            time_str = f"{hours:02}:{minutes:02}"
            
            # Obter data de conclusão
            history = task.get('history', [])
            last_event = history[-1] if history else '-'
            
            self.table.setItem(row, 0, QTableWidgetItem(task.get('ticket', '-')))
            self.table.setItem(row, 1, QTableWidgetItem(task.get('email', '-')))
            self.table.setItem(row, 2, QTableWidgetItem(task.get('category', '-')))
            self.table.setItem(row, 3, QTableWidgetItem(task.get('priority', '-')))
            self.table.setItem(row, 4, QTableWidgetItem(last_event))
            self.table.setItem(row, 5, QTableWidgetItem(time_str))
            self.table.setItem(row, 6, QTableWidgetItem(task.get('description', '-')[:50]))
    
    def export_excel(self):
        """Exporta logs para Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exportar Excel", "", "Excel Files (*.xlsx)"
            )
            
            if not filename:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Logs de Atividades"
            
            headers = ["Ticket", "Email", "Categoria", "Descrição", "Prioridade", 
                      "Status", "Tempo Total (s)", "Tempo Formatado", "Histórico"]
            ws.append(headers)
            
            # Estilo do header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Dados
            tasks = self.db.get_completed_tasks()
            for task in tasks.values():
                total_seconds = task.get('total_seconds', 0)
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                time_str = f"{hours:02}:{minutes:02}"
                
                ws.append([
                    task.get('ticket', ''),
                    task.get('email', ''),
                    task.get('category', ''),
                    task.get('description', ''),
                    task.get('priority', ''),
                    task.get('status', ''),
                    total_seconds,
                    time_str,
                    "\n".join(task.get('history', []))
                ])
            
            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filename)
            QMessageBox.information(self, "Sucesso", "Log exportado com sucesso!")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar: {str(e)}")


class MainWindow(QMainWindow):
    """Janela principal do sistema"""
    
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.active_tasks = set()
        self.notification_timers = {}
        self.last_notification_time = {}
        self.notification_interval = 1800  # 30 minutos em segundos
        self.multi_task_notified = set()  # Controle para não notificar repetido
        
        self.setup_ui()
        self.setup_tray()
        self.setup_timer()
        self.refresh_table()
        self.recover_active_tasks()
    
    def setup_ui(self):
        """Configura a interface principal"""
        self.setWindowTitle("Gestão de Atividades - TI")
        self.setMinimumSize(1100, 700)
        
        # Tema escuro
        self.setStyleSheet("""
            QMainWindow {
                background-color: #2b2b2b;
            }
            QWidget {
                background-color: #2b2b2b;
                color: #ffffff;
                font-family: Arial;
                font-size: 13px;
            }
            QLabel {
                color: #ffffff;
                padding: 3px;
            }
            QLineEdit, QComboBox, QDateEdit {
                background-color: #3c3f41;
                color: #ffffff;
                border: 1px solid #555555;
                border-radius: 4px;
                padding: 6px;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
                border: 1px solid #4a90e2;
            }
            QPushButton {
                background-color: #4a90e2;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #357abd;
            }
            QPushButton:pressed {
                background-color: #2a5f8f;
            }
            QPushButton:disabled {
                background-color: #555555;
                color: #888888;
            }
            QTableWidget {
                background-color: #323232;
                color: #ffffff;
                border: 1px solid #555555;
                gridline-color: #444444;
            }
            QTableWidget::item:selected {
                background-color: #505050;
            }
            QHeaderView::section {
                background-color: #1e1e1e;
                color: #ffffff;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
            QGroupBox {
                border: 1px solid #555555;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
                font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QScrollBar:vertical {
                background-color: #2b2b2b;
                width: 12px;
            }
            QScrollBar::handle:vertical {
                background-color: #555555;
                border-radius: 6px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
        """)
        
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # === Seção de Nova Demanda ===
        new_task_group = QGroupBox("📝 Nova Demanda")
        new_task_layout = QVBoxLayout(new_task_group)
        
        # Campo de entrada do chamado
        input_layout = QHBoxLayout()
        input_layout.addWidget(QLabel("Cole a linha do chamado:"))
        self.call_input = QLineEdit()
        self.call_input.setPlaceholderText("Cole aqui a linha completa do chamado...")
        self.call_input.textChanged.connect(self.parse_call)
        input_layout.addWidget(self.call_input)
        
        btn_clear = QPushButton("🗑️ Limpar")
        btn_clear.setMaximumWidth(100)
        btn_clear.clicked.connect(self.clear_form)
        input_layout.addWidget(btn_clear)
        
        new_task_layout.addLayout(input_layout)
        
        # Dados extraídos
        info_layout = QHBoxLayout()
        self.lbl_ticket = QLabel("Ticket: -")
        self.lbl_ticket.setStyleSheet("color: #4a90e2; font-weight: bold;")
        info_layout.addWidget(self.lbl_ticket)
        
        self.lbl_email = QLabel("Email: -")
        info_layout.addWidget(self.lbl_email)
        
        self.lbl_category = QLabel("Categoria: -")
        info_layout.addWidget(self.lbl_category)
        
        info_layout.addStretch()
        new_task_layout.addLayout(info_layout)
        
        # Prioridade e botão adicionar
        action_layout = QHBoxLayout()
        action_layout.addWidget(QLabel("Prioridade:"))
        self.priority_combo = QComboBox()
        self.priority_combo.addItems(["Baixa", "Média", "Alta"])
        self.priority_combo.setCurrentIndex(1)
        self.priority_combo.setMaximumWidth(150)
        action_layout.addWidget(self.priority_combo)
        
        btn_add = QPushButton("➕ Adicionar Tarefa")
        btn_add.clicked.connect(self.add_task)
        action_layout.addWidget(btn_add)
        
        action_layout.addStretch()
        new_task_layout.addLayout(action_layout)
        
        main_layout.addWidget(new_task_group)
        
        # === Tabela de Tarefas ===
        table_group = QGroupBox("📋 Tarefas")
        table_layout = QVBoxLayout(table_group)
        
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "ID", "Ticket", "Descrição", "Prioridade", "Status", "Tempo", "Ações"
        ])
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(0, 80)   # ID
        self.table.setColumnWidth(1, 120)  # Ticket
        self.table.setColumnWidth(2, 350)  # Descrição
        self.table.setColumnWidth(3, 80)   # Prioridade
        self.table.setColumnWidth(4, 100)  # Status
        self.table.setColumnWidth(5, 90)   # Tempo
        self.table.setColumnWidth(6, 200)  # Ações
        
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        self.table.cellDoubleClicked.connect(self.on_double_click)
        
        table_layout.addWidget(self.table)
        main_layout.addWidget(table_group, stretch=1)
        
        # === Botões de Ação ===
        buttons_layout = QHBoxLayout()
        
        self.btn_start = QPushButton("▶️ Iniciar/Retomar")
        self.btn_start.clicked.connect(self.start_task)
        self.btn_start.setEnabled(False)
        buttons_layout.addWidget(self.btn_start)
        
        self.btn_pause = QPushButton("⏸️ Pausar")
        self.btn_pause.clicked.connect(self.pause_task)
        self.btn_pause.setEnabled(False)
        buttons_layout.addWidget(self.btn_pause)
        
        self.btn_finish = QPushButton("✅ Concluir")
        self.btn_finish.clicked.connect(self.finish_task)
        self.btn_finish.setEnabled(False)
        buttons_layout.addWidget(self.btn_finish)
        
        self.btn_edit = QPushButton("✏️ Editar")
        self.btn_edit.clicked.connect(self.edit_task)
        self.btn_edit.setEnabled(False)
        buttons_layout.addWidget(self.btn_edit)
        
        buttons_layout.addStretch()
        
        self.btn_logs = QPushButton("📊 Ver Logs")
        self.btn_logs.clicked.connect(self.open_logs)
        buttons_layout.addWidget(self.btn_logs)
        
        main_layout.addLayout(buttons_layout)
        
        # Variáveis de controle
        self.parsed_data = None
        self.selected_row = -1
    
    def setup_tray(self):
        """Configura ícone na bandeja do sistema"""
        self.tray_icon = QSystemTrayIcon(self)
        
        # Criar ícone simples (pode ser substituído por um arquivo .ico)
        pixmap = QPixmap(64, 64)
        pixmap.fill(QColor("#4a90e2"))
        icon = QIcon(pixmap)
        
        self.tray_icon.setIcon(icon)
        self.tray_icon.setToolTip("Gestão de Atividades")
        
        # Menu do tray
        tray_menu = QMenu()
        
        restore_action = QAction("Restaurar", self)
        restore_action.triggered.connect(self.restore_from_tray)
        tray_menu.addAction(restore_action)
        
        logs_action = QAction("Ver Logs", self)
        logs_action.triggered.connect(self.open_logs)
        tray_menu.addAction(logs_action)
        
        tray_menu.addSeparator()
        
        quit_action = QAction("Sair", self)
        quit_action.triggered.connect(self.quit_app)
        tray_menu.addAction(quit_action)
        
        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.activated.connect(self.tray_activated)
        self.tray_icon.show()
    
    def tray_activated(self, reason):
        """Handle de ativação do tray"""
        if reason == QSystemTrayIcon.ActivationReason.DoubleClick:
            self.restore_from_tray()
    
    def restore_from_tray(self):
        """Restaura janela do tray"""
        self.showNormal()
        self.activateWindow()
        self.raise_()
    
    def closeEvent(self, event):
        """Minimiza para tray ao invés de fechar"""
        event.ignore()
        self.hide()
        
        if HAS_PLYER:
            try:
                notification.notify(
                    title="Gestão de Atividades",
                    message="Sistema minimizado na bandeja.",
                    timeout=3
                )
            except:
                pass
    
    def quit_app(self):
        """Fecha o aplicativo completamente"""
        # Salvar estado das tarefas ativas
        for task_id in list(self.active_tasks):
            task = self.db.get_task(task_id)
            if task and task.get('status') == 'Em Andamento':
                # Pausar automaticamente ao fechar
                now = datetime.now()
                if 'last_start_time' in task:
                    last_start = datetime.strptime(task['last_start_time'], "%Y-%m-%d %H:%M:%S")
                    delta = (now - last_start).total_seconds()
                    task['total_seconds'] = task.get('total_seconds', 0) + delta
                
                task['status'] = 'Pausado'
                task['history'].append(f"Sistema fechado (pausa automática): {now.strftime('%d/%m %H:%M')}")
                self.db.update_task(task_id, task)
        
        self.tray_icon.hide()
        QApplication.quit()
    
    def setup_timer(self):
        """Configura timer para atualizações e notificações"""
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_timers)
        self.timer.start(1000)  # Atualiza a cada segundo
    
    def update_timers(self):
        """Atualiza timers e verifica notificações"""
        # Atualizar visualização do tempo das tarefas ativas
        if self.active_tasks:
            self.refresh_table()
        
        # Verificar múltiplas tarefas ativas
        self.check_multiple_tasks()
    
    def check_multiple_tasks(self):
        """Verifica se há múltiplas tarefas ativas e notifica"""
        if len(self.active_tasks) <= 1:
            # Resetar notificações quando voltar para uma tarefa
            self.multi_task_notified.clear()
            return
        
        now = time.time()
        
        # Notificar apenas uma vez por ciclo de 30 min para todas as tarefas
        for task_id in self.active_tasks:
            last_notif = self.last_notification_time.get(task_id, 0)
            
            if now - last_notif >= self.notification_interval:
                task = self.db.get_task(task_id)
                if task:
                    # Criar mensagem listando todas as tarefas ativas
                    active_list = []
                    for tid in self.active_tasks:
                        t = self.db.get_task(tid)
                        if t:
                            active_list.append(f"• {t.get('ticket', '?')}: {t.get('description', '')[:30]}")
                    
                    msg = (f"⚠️ Você tem {len(self.active_tasks)} tarefas em andamento simultâneo!\n\n"
                           f"Tarefas ativas:\n{chr(10).join(active_list)}\n\n"
                           f"Lembre-se de gerenciar seu tempo!")
                    
                    # Notificação via system tray
                    self.tray_icon.showMessage(
                        "Gestão de Atividades - Múltiplas Demandas",
                        msg,
                        QSystemTrayIcon.MessageIcon.Warning,
                        15000
                    )
                    
                    # Notificação via plyer (se disponível)
                    if HAS_PLYER:
                        try:
                            notification.notify(
                                title="Gestão de Atividades - Múltiplas Tarefas",
                                message=msg,
                                timeout=15
                            )
                        except:
                            pass
                    
                    # Atualizar tempo da notificação para TODAS as tarefas ativas
                    for tid in self.active_tasks:
                        self.last_notification_time[tid] = now
                    break  # Sai após notificar uma vez
    
    def parse_call(self, text):
        """Parse automático da linha do chamado"""
        if not text:
            self.parsed_data = None
            self.lbl_ticket.setText("Ticket: -")
            self.lbl_email.setText("Email: -")
            self.lbl_category.setText("Categoria: -")
            return
        
        result = TaskParser.parse(text)
        
        if result:
            self.parsed_data = result
            self.lbl_ticket.setText(f"Ticket: {result['ticket']}")
            self.lbl_email.setText(f"Email: {result['email']}")
            self.lbl_category.setText(f"Categoria: {result['category']}")
        else:
            self.parsed_data = None
            self.lbl_ticket.setText("Ticket: -")
            self.lbl_email.setText("Email: -")
            self.lbl_category.setText("Categoria: -")
    
    def clear_form(self):
        """Limpa formulário"""
        self.call_input.clear()
        self.priority_combo.setCurrentIndex(1)
        self.parsed_data = None
        self.lbl_ticket.setText("Ticket: -")
        self.lbl_email.setText("Email: -")
        self.lbl_category.setText("Categoria: -")
    
    def add_task(self):
        """Adiciona nova tarefa"""
        if not self.parsed_data:
            QMessageBox.warning(self, "Atenção", 
                              "Cole a linha do chamado e aguarde a extração dos dados.")
            return
        
        task_id = datetime.now().strftime("%Y%m%d%H%M%S")
        
        task = {
            'id': task_id,
            'ticket': self.parsed_data['ticket'],
            'email': self.parsed_data['email'],
            'category': self.parsed_data['category'],
            'description': self.parsed_data['description'],
            'priority': self.priority_combo.currentText(),
            'status': 'Aguardando',
            'start_time': None,
            'end_time': None,
            'total_seconds': 0,
            'history': [],
            'last_start_time': None
        }
        
        self.db.add_task(task)
        self.refresh_table()
        self.clear_form()
        
        QMessageBox.information(self, "Sucesso", "Tarefa adicionada com sucesso!")
    
    def refresh_table(self, show_completed=False):
        """Atualiza tabela de tarefas"""
        self.table.setRowCount(0)
        
        tasks = self.db.get_all_tasks()
        
        # Cores por prioridade
        priority_colors = {
            'Baixa': '#4caf50',
            'Média': '#2196f3',
            'Alta': '#ff9800'
        }
        
        # Salvar seleção atual
        selected_task_id = None
        if self.selected_row >= 0 and self.selected_row < self.table.rowCount():
            selected_task_id = self.get_task_id_from_row(self.selected_row)
        
        for task_id, task in tasks.items():
            # Filtrar tarefas concluídas (a menos que show_completed seja True)
            if task.get('status') == 'Concluido' and not show_completed:
                continue
            
            row = self.table.rowCount()
            self.table.insertRow(row)
            
            # Formatar tempo
            total_seconds = task.get('total_seconds', 0)
            
            # Se estiver em andamento, calcular tempo adicional
            if task.get('status') == 'Em Andamento' and task.get('last_start_time'):
                try:
                    last_start = datetime.strptime(task['last_start_time'], "%Y-%m-%d %H:%M:%S")
                    delta = (datetime.now() - last_start).total_seconds()
                    total_seconds += delta
                except:
                    pass
            
            hours = int(total_seconds // 3600)
            minutes = int((total_seconds % 3600) // 60)
            seconds = int(total_seconds % 60)
            time_str = f"{hours:02}:{minutes:02}:{seconds:02}"
            
            # Definir itens
            self.table.setItem(row, 0, QTableWidgetItem(task_id[-4:]))
            self.table.setItem(row, 1, QTableWidgetItem(task.get('ticket', '-')))
            self.table.setItem(row, 2, QTableWidgetItem(task.get('description', '-')[:50]))
            
            # Prioridade com cor
            priority_item = QTableWidgetItem(task.get('priority', '-'))
            priority_color = priority_colors.get(task.get('priority'), '#888888')
            priority_item.setBackground(QColor(priority_color))
            priority_item.setForeground(QColor('#ffffff'))
            self.table.setItem(row, 3, priority_item)
            
            # Status
            status_item = QTableWidgetItem(task.get('status', '-'))
            if task.get('status') == 'Em Andamento':
                status_item.setBackground(QColor('#4caf50'))
                status_item.setForeground(QColor('#ffffff'))
            elif task.get('status') == 'Pausado':
                status_item.setBackground(QColor('#ff9800'))
                status_item.setForeground(QColor('#ffffff'))
            elif task.get('status') == 'Concluido':
                status_item.setBackground(QColor('#666666'))
                status_item.setForeground(QColor('#ffffff'))
            self.table.setItem(row, 4, status_item)
            
            # Tempo
            self.table.setItem(row, 5, QTableWidgetItem(time_str))
            
            # Ações (botões na célula)
            actions_widget = QWidget()
            actions_layout = QHBoxLayout(actions_widget)
            actions_layout.setContentsMargins(5, 2, 5, 2)
            actions_layout.setSpacing(5)
            
            if task.get('status') in ['Aguardando', 'Pausado']:
                btn = QPushButton("▶️")
                btn.setMaximumWidth(40)
                btn.clicked.connect(lambda checked, tid=task_id: self.quick_start(tid))
                actions_layout.addWidget(btn)
            elif task.get('status') == 'Em Andamento':
                btn = QPushButton("⏸️")
                btn.setMaximumWidth(40)
                btn.clicked.connect(lambda checked, tid=task_id: self.quick_pause(tid))
                actions_layout.addWidget(btn)
            
            if task.get('status') in ['Em Andamento', 'Pausado']:
                btn = QPushButton("✅")
                btn.setMaximumWidth(40)
                btn.clicked.connect(lambda checked, tid=task_id: self.quick_finish(tid))
                actions_layout.addWidget(btn)
            
            self.table.setCellWidget(row, 6, actions_widget)
    
    def on_selection_changed(self):
        """Handle seleção de tarefa"""
        selected_rows = self.table.selectedItems()
        if selected_rows and len(selected_rows) > 0:
            # Manter a seleção sem perder o foco
            current_row = self.selected_row
            if current_row >= 0 and current_row < self.table.rowCount():
                self.selected_row = current_row
            else:
                self.selected_row = selected_rows[0].row()
            
            task_id = self.get_task_id_from_row(self.selected_row)
            
            if task_id:
                task = self.db.get_task(task_id)
                if task:
                    status = task.get('status')
                    self.btn_start.setEnabled(status in ['Aguardando', 'Pausado'])
                    self.btn_pause.setEnabled(status == 'Em Andamento')
                    self.btn_finish.setEnabled(status in ['Em Andamento', 'Pausado'])
                    self.btn_edit.setEnabled(True)  # Habilitar edição para qualquer tarefa ativa
        else:
            self.selected_row = -1
            self.btn_start.setEnabled(False)
            self.btn_pause.setEnabled(False)
            self.btn_finish.setEnabled(False)
            self.btn_edit.setEnabled(False)
    
    def on_double_click(self, row, column):
        """Duplo clique inicia/pausa tarefa"""
        task_id = self.get_task_id_from_row(row)
        if task_id:
            task = self.db.get_task(task_id)
            if task:
                if task.get('status') in ['Aguardando', 'Pausado']:
                    self.quick_start(task_id)
                elif task.get('status') == 'Em Andamento':
                    self.quick_pause(task_id)
    
    def get_task_id_from_row(self, row):
        """Obtém ID da tarefa a partir da linha"""
        if row >= 0 and row < self.table.rowCount():
            item = self.table.item(row, 0)
            if item:
                short_id = item.text()
                # Buscar tarefa completa
                for task_id, task in self.db.get_all_tasks().items():
                    if task_id.endswith(short_id):
                        return task_id
        return None
    
    def get_selected_task_id(self):
        """Obtém ID da tarefa selecionada"""
        if self.selected_row >= 0:
            return self.get_task_id_from_row(self.selected_row)
        return None
    
    def start_task(self, task_id=None):
        """Inicia ou retoma tarefa"""
        if task_id is None:
            task_id = self.get_selected_task_id()
        
        if not task_id:
            return
        
        task = self.db.get_task(task_id)
        if not task:
            return
        
        now = datetime.now()
        now_str = now.strftime("%Y-%m-%d %H:%M:%S")
        
        if task.get('status') == 'Aguardando':
            task['start_time'] = now_str
            task['history'].append(f"Início: {now.strftime('%d/%m %H:%M')}")
        elif task.get('status') == 'Pausado':
            task['history'].append(f"Retomado: {now.strftime('%d/%m %H:%M')}")
        
        task['status'] = 'Em Andamento'
        task['last_start_time'] = now_str
        self.active_tasks.add(task_id)
        
        self.db.update_task(task_id, task)
        self.refresh_table()
        self.on_selection_changed()
        self.check_multiple_tasks()
    
    def quick_start(self, task_id):
        """Início rápido via botão na tabela"""
        self.start_task(task_id)
    
    def pause_task(self, task_id=None):
        """Pausa tarefa"""
        if task_id is None:
            task_id = self.get_selected_task_id()
        
        if not task_id:
            return
        
        task = self.db.get_task(task_id)
        if not task or task.get('status') != 'Em Andamento':
            return
        
        now = datetime.now()
        
        # Calcular tempo desde último início
        if task.get('last_start_time'):
            try:
                last_start = datetime.strptime(task['last_start_time'], "%Y-%m-%d %H:%M:%S")
                delta = (now - last_start).total_seconds()
                task['total_seconds'] = task.get('total_seconds', 0) + delta
            except:
                pass
        
        task['status'] = 'Pausado'
        task['last_start_time'] = None
        task['history'].append(f"Pausa: {now.strftime('%d/%m %H:%M')} (Total: {self.format_time(task['total_seconds'])})")
        
        if task_id in self.active_tasks:
            self.active_tasks.remove(task_id)
        
        self.db.update_task(task_id, task)
        self.refresh_table()
        self.on_selection_changed()
    
    def quick_pause(self, task_id):
        """Pausa rápida via botão na tabela"""
        self.pause_task(task_id)
    
    def finish_task(self, task_id=None):
        """Conclui tarefa"""
        if task_id is None:
            task_id = self.get_selected_task_id()
        
        if not task_id:
            return
        
        task = self.db.get_task(task_id)
        if not task:
            return
        
        now = datetime.now()
        
        # Calcular tempo final se estiver em andamento
        if task.get('status') == 'Em Andamento' and task.get('last_start_time'):
            try:
                last_start = datetime.strptime(task['last_start_time'], "%Y-%m-%d %H:%M:%S")
                delta = (now - last_start).total_seconds()
                task['total_seconds'] = task.get('total_seconds', 0) + delta
            except:
                pass
        
        task['status'] = 'Concluido'
        task['end_time'] = now.strftime("%Y-%m-%d %H:%M:%S")
        task['last_start_time'] = None
        task['history'].append(f"Concluído: {now.strftime('%d/%m %H:%M')}")
        
        if task_id in self.active_tasks:
            self.active_tasks.remove(task_id)
        
        self.db.update_task(task_id, task)
        self.refresh_table()
        self.on_selection_changed()
        
        QMessageBox.information(
            self, "Concluído",
            f"Tarefa {task.get('ticket', '?')} finalizada!\n"
            f"Tempo total: {self.format_time(task['total_seconds'])}"
        )
    
    def quick_finish(self, task_id):
        """Conclusão rápida via botão na tabela"""
        self.finish_task(task_id)
    
    def edit_task(self, task_id=None):
        """Edita tarefa selecionada"""
        # Se não passou ID, tenta pegar da seleção atual
        if task_id is None:
            selected_items = self.table.selectedItems()
            if not selected_items:
                QMessageBox.warning(self, "Editar", "Selecione uma tarefa válida para editar.")
                return
            
            # Pega a primeira coluna da primeira linha selecionada para garantir o ID
            row = selected_items[0].row()
            ticket_item = self.table.item(row, 0)
            
            if not ticket_item:
                QMessageBox.warning(self, "Editar", "Erro ao identificar a tarefa.")
                return
                
            task_id = ticket_item.text()
        
        # Verifica se o ID existe no banco
        if not task_id or task_id not in self.db.tasks:
            QMessageBox.warning(self, "Editar", "Tarefa selecionada não encontrada ou já foi concluída.")
            return
        
        task = self.db.tasks[task_id]
        
        # Criar diálogo de edição
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Editar Tarefa {task_id}")
        dialog.setModal(True)
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout()
        
        # Título
        layout.addWidget(QLabel("Título da Tarefa:"))
        txt_title = QLineEdit(task.get('title', ''))
        layout.addWidget(txt_title)
        
        # Prioridade
        layout.addWidget(QLabel("Prioridade:"))
        cmb_priority = QComboBox()
        cmb_priority.addItems(["Baixa", "Média", "Alta"])
        current_priority = task.get('priority', 'Média')
        cmb_priority.setCurrentText(current_priority)
        layout.addWidget(cmb_priority)
        
        # Descrição
        layout.addWidget(QLabel("Descrição:"))
        txt_description = QTextEdit()  # Mudado para QTextEdit para permitir múltiplas linhas
        txt_description.setPlainText(task.get('description', ''))
        txt_description.setMaximumHeight(150)
        layout.addWidget(txt_description)
        
        # Botões
        btn_layout = QHBoxLayout()
        btn_save = QPushButton("💾 Salvar")
        btn_cancel = QPushButton("❌ Cancelar")
        btn_layout.addWidget(btn_save)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)
        
        dialog.setLayout(layout)
        
        # Conectar botões
        btn_save.clicked.connect(dialog.accept)
        btn_cancel.clicked.connect(dialog.reject)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_title = txt_title.text().strip()
            if not new_title:
                QMessageBox.warning(self, "Erro", "O título não pode estar vazio.")
                return

            # Atualizar tarefa
            task['title'] = new_title
            task['priority'] = cmb_priority.currentText()
            task['description'] = txt_description.toPlainText().strip()
            
            # Adicionar histórico de edição
            task['history'].append(f"Tarefa editada em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            
            self.db.update_task(task_id, task)
            self.refresh_table()
            self.on_selection_changed()
            QMessageBox.information(self, "Editar", "Tarefa atualizada com sucesso!")
    
    def format_time(self, seconds):
        """Formata segundos em HH:MM:SS"""
        h = int(seconds // 3600)
        m = int((seconds % 3600) // 60)
        s = int(seconds % 60)
        return f"{h:02}:{m:02}:{s:02}"
    
    def recover_active_tasks(self):
        """Recupera tarefas que estavam ativas ao fechar"""
        for task_id, task in self.db.get_all_tasks().items():
            if task.get('status') == 'Em Andamento':
                # Marcar como pausado para segurança
                task['status'] = 'Pausado'
                task['history'].append("Reinicialização do sistema (ajuste necessário)")
                self.db.update_task(task_id, task)
        
        self.refresh_table()
    
    def open_logs(self):
        """Abre janela de logs"""
        dialog = LogsDialog(self.db, self)
        dialog.exec()


def setup_startup():
    """Configura para iniciar com o Windows"""
    if sys.platform == 'win32':
        try:
            import winreg
            
            key_path = r"Software\Microsoft\Windows\CurrentVersion\Run"
            app_name = "GestaoAtividades"
            app_path = os.path.abspath(sys.argv[0])
            
            # Se for executável compilado
            if app_path.endswith('.exe'):
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path, 0, winreg.KEY_SET_VALUE)
                winreg.SetValueEx(key, app_name, 0, winreg.REG_SZ, app_path)
                winreg.CloseKey(key)
                print("Startup configurado com sucesso!")
        except Exception as e:
            print(f"Erro ao configurar startup: {e}")


def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Gestão de Atividades")
    
    # Configurar startup (apenas se solicitado ou primeira execução)
    # setup_startup()  # Descomentar se quiser ativar automaticamente
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
