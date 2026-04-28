"""
Gestão de Atividades - Sistema de Gerenciamento de Tarefas para TI
Versão: 1.0.0
Autor: Isaacssjr
Descrição: Sistema completo para gerenciamento de demandas de TI com controle de tempo,
           notificações, logs e interface moderna em tema escuro.
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from datetime import datetime, timedelta
import json
import os
import sys
import threading
import time
from collections import defaultdict
from pathlib import Path

# Tentar importar plyer para notificações nativas do Windows
try:
    from plyer import notification
    PLYER_AVAILABLE = True
except ImportError:
    PLYER_AVAILABLE = False

# Configurações do sistema
CONFIG_DIR = Path(os.path.expanduser("~")) / "GestaoDeAtividades"
CONFIG_FILE = CONFIG_DIR / "config.json"
LOGS_FILE = CONFIG_DIR / "logs.json"
TASKS_FILE = CONFIG_DIR / "tasks.json"

# Garantir que o diretório de configuração existe
CONFIG_DIR.mkdir(exist_ok=True)


class Task:
    """Classe que representa uma tarefa/chamado"""
    
    def __init__(self, ticket, email, category, description, priority="media"):
        self.ticket = ticket
        self.email = email
        self.category = category
        self.description = description
        self.priority = priority
        self.status = "pendente"  # pendente, em_andamento, pausado, concluido
        self.created_at = datetime.now()
        self.started_at = None
        self.finished_at = None
        self.total_time = timedelta(0)
        self.pause_start = None
        self.time_sessions = []  # Lista de tuplas (inicio, fim)
        self.current_session_start = None
        
    def start(self):
        """Inicia ou retoma a tarefa"""
        if self.status == "pendente":
            self.started_at = datetime.now()
            self.current_session_start = self.started_at
        elif self.status == "pausado":
            self.current_session_start = datetime.now()
        
        self.status = "em_andamento"
        
    def pause(self):
        """Pausa a tarefa"""
        if self.status == "em_andamento" and self.current_session_start:
            pause_time = datetime.now()
            session_duration = pause_time - self.current_session_start
            self.time_sessions.append((self.current_session_start, pause_time))
            self.total_time += session_duration
            self.current_session_start = None
            self.status = "pausado"
            
    def resume(self):
        """Retoma a tarefa pausada"""
        if self.status == "pausado":
            self.current_session_start = datetime.now()
            self.status = "em_andamento"
            
    def finish(self):
        """Conclui a tarefa"""
        if self.current_session_start:
            finish_time = datetime.now()
            session_duration = finish_time - self.current_session_start
            self.time_sessions.append((self.current_session_start, finish_time))
            self.total_time += session_duration
            self.current_session_start = None
            
        self.finished_at = datetime.now()
        self.status = "concluido"
        
    def get_total_time_str(self):
        """Retorna o tempo total formatado como string"""
        total = self.total_time
        hours = int(total.total_seconds() // 3600)
        minutes = int((total.total_seconds() % 3600) // 60)
        seconds = int(total.total_seconds() % 60)
        
        if hours > 0:
            return f"{hours}h {minutes}m {seconds}s"
        elif minutes > 0:
            return f"{minutes}m {seconds}s"
        else:
            return f"{seconds}s"
            
    def to_dict(self):
        """Converte a tarefa para dicionário (para serialização JSON)"""
        return {
            'ticket': self.ticket,
            'email': self.email,
            'category': self.category,
            'description': self.description,
            'priority': self.priority,
            'status': self.status,
            'created_at': self.created_at.isoformat(),
            'started_at': self.started_at.isoformat() if self.started_at else None,
            'finished_at': self.finished_at.isoformat() if self.finished_at else None,
            'total_time_seconds': self.total_time.total_seconds(),
            'time_sessions': [(s.isoformat(), e.isoformat()) for s, e in self.time_sessions],
            'current_session_start': self.current_session_start.isoformat() if self.current_session_start else None
        }
        
    @classmethod
    def from_dict(cls, data):
        """Cria uma tarefa a partir de um dicionário"""
        task = cls(
            data['ticket'],
            data['email'],
            data['category'],
            data['description'],
            data['priority']
        )
        task.status = data['status']
        task.created_at = datetime.fromisoformat(data['created_at'])
        task.started_at = datetime.fromisoformat(data['started_at']) if data['started_at'] else None
        task.finished_at = datetime.fromisoformat(data['finished_at']) if data['finished_at'] else None
        task.total_time = timedelta(seconds=data['total_time_seconds'])
        task.time_sessions = [
            (datetime.fromisoformat(s), datetime.fromisoformat(e)) 
            for s, e in data.get('time_sessions', [])
        ]
        if data.get('current_session_start'):
            task.current_session_start = datetime.fromisoformat(data['current_session_start'])
            
        return task


class NotificationManager:
    """Gerencia notificações do sistema"""
    
    def __init__(self):
        self.last_notification = {}
        self.notification_interval = 30 * 60  # 30 minutos em segundos
        
    def send_notification(self, title, message):
        """Envia uma notificação para o usuário"""
        if PLYER_AVAILABLE:
            try:
                notification.notify(
                    title=title,
                    message=message,
                    app_name="Gestão de Atividades",
                    timeout=10
                )
            except Exception as e:
                print(f"Erro ao enviar notificação: {e}")
                self._show_fallback_notification(title, message)
        else:
            self._show_fallback_notification(title, message)
            
    def _show_fallback_notification(self, title, message):
        """Mostra notificação fallback usando Tkinter"""
        # Esta função será chamada pelo main thread através de after
        pass
        
    def check_multiple_tasks(self, active_tasks):
        """Verifica se há múltiplas tarefas ativas e notifica se necessário"""
        if len(active_tasks) < 2:
            return
            
        now = datetime.now()
        
        for task in active_tasks:
            task_key = task.ticket
            
            # Verifica se já notificou sobre esta tarefa recentemente
            if task_key in self.last_notification:
                last_notif = self.last_notification[task_key]
                if (now - last_notif).total_seconds() < self.notification_interval:
                    continue
                    
            # Notifica sobre as outras tarefas
            other_tasks = [t for t in active_tasks if t.ticket != task.ticket]
            if other_tasks:
                other_names = ", ".join([f"{t.ticket}" for t in other_tasks])
                self.send_notification(
                    "⚠️ Múltiplas Tarefas em Andamento",
                    f"Você está atendendo o chamado {task.ticket}.\n\n"
                    f"Lembrete: Você também tem {len(other_tasks)} outra(s) tarefa(s) em andamento:\n{other_names}\n\n"
                    f"Tempo desde o último lembrete: 30 minutos"
                )
                self.last_notification[task_key] = now


class ThemeManager:
    """Gerencia o tema escuro da aplicação"""
    
    COLORS = {
        'bg_dark': '#1e1e1e',
        'bg_medium': '#2d2d2d',
        'bg_light': '#3d3d3d',
        'fg_primary': '#ffffff',
        'fg_secondary': '#b0b0b0',
        'accent': '#0078d4',
        'accent_hover': '#1084d8',
        'success': '#4caf50',
        'warning': '#ff9800',
        'danger': '#f44336',
        'info': '#2196f3',
        'priority_baixa': '#4caf50',
        'priority_media': '#2196f3',
        'priority_alta': '#ff9800',
        'priority_urgente': '#f44336'
    }
    
    @classmethod
    def get_priority_color(cls, priority):
        """Retorna a cor baseada na prioridade"""
        colors = {
            'baixa': cls.COLORS['priority_baixa'],
            'media': cls.COLORS['priority_media'],
            'alta': cls.COLORS['priority_alta'],
            'urgente': cls.COLORS['priority_urgente']
        }
        return colors.get(priority, cls.COLORS['priority_media'])
    
    @classmethod
    def get_priority_name(cls, priority):
        """Retorna o nome legível da prioridade"""
        names = {
            'baixa': 'Baixa',
            'media': 'Média',
            'alta': 'Alta',
            'urgente': 'Urgente'
        }
        return names.get(priority, 'Média')


class LogsWindow(tk.Toplevel):
    """Janela de visualização e filtragem de logs"""
    
    def __init__(self, parent, task_manager):
        super().__init__(parent)
        self.task_manager = task_manager
        self.title("📊 Histórico de Chamados - Gestão de Atividades")
        self.geometry("1000x600")
        self.minsize(800, 500)
        
        self.configure(bg=ThemeManager.COLORS['bg_dark'])
        
        # Estilo
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self._configure_styles()
        
        self._create_widgets()
        self._load_logs()
        
    def _configure_styles(self):
        """Configura os estilos dos widgets"""
        colors = ThemeManager.COLORS
        
        self.style.configure("Treeview",
                            background=colors['bg_medium'],
                            foreground=colors['fg_primary'],
                            fieldbackground=colors['bg_medium'],
                            rowheight=25,
                            font=('Segoe UI', 10))
        
        self.style.configure("Treeview.Heading",
                            background=colors['bg_light'],
                            foreground=colors['fg_primary'],
                            font=('Segoe UI', 10, 'bold'))
        
        self.style.map("Treeview",
                      background=[('selected', colors['accent'])],
                      foreground=[('selected', colors['fg_primary'])])
        
        self.style.configure("TButton",
                            background=colors['accent'],
                            foreground=colors['fg_primary'],
                            font=('Segoe UI', 10))
        
        self.style.configure("TLabel",
                            background=colors['bg_dark'],
                            foreground=colors['fg_primary'],
                            font=('Segoe UI', 10))
        
    def _create_widgets(self):
        """Cria os widgets da janela"""
        colors = ThemeManager.COLORS
        
        # Frame principal
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Frame de filtros
        filter_frame = ttk.LabelFrame(main_frame, text="🔍 Filtros", padding="10")
        filter_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Filtro por texto
        ttk.Label(filter_frame, text="Buscar:").grid(row=0, column=0, padx=(0, 5), sticky=tk.W)
        self.text_filter = ttk.Entry(filter_frame, width=40)
        self.text_filter.grid(row=0, column=1, padx=(0, 10), sticky=tk.W)
        self.text_filter.bind('<KeyRelease>', lambda e: self._apply_filters())
        
        # Filtro por data
        ttk.Label(filter_frame, text="De:").grid(row=0, column=2, padx=(10, 5), sticky=tk.W)
        self.date_from = ttk.Entry(filter_frame, width=12)
        self.date_from.grid(row=0, column=3, padx=(0, 5), sticky=tk.W)
        self.date_from.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        ttk.Label(filter_frame, text="Até:").grid(row=0, column=4, padx=(10, 5), sticky=tk.W)
        self.date_to = ttk.Entry(filter_frame, width=12)
        self.date_to.grid(row=0, column=5, padx=(0, 10), sticky=tk.W)
        self.date_to.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        # Botões de filtro
        btn_frame = ttk.Frame(filter_frame)
        btn_frame.grid(row=0, column=6, padx=(10, 0))
        
        ttk.Button(btn_frame, text="Aplicar", command=self._apply_filters).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_frame, text="Limpar", command=self._clear_filters).pack(side=tk.LEFT)
        
        # Botão de exportar
        ttk.Button(btn_frame, text="📥 Exportar Excel", command=self._export_to_excel).pack(side=tk.LEFT, padx=(10, 0))
        
        # Treeview para os logs
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ('ticket', 'email', 'category', 'priority', 'status', 'start', 'end', 'duration')
        
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=20)
        
        # Configurar colunas
        self.tree.heading('ticket', text='Ticket')
        self.tree.column('ticket', width=120)
        
        self.tree.heading('email', text='E-mail')
        self.tree.column('email', width=200)
        
        self.tree.heading('category', text='Categoria')
        self.tree.column('category', width=250)
        
        self.tree.heading('priority', text='Prioridade')
        self.tree.column('priority', width=80)
        
        self.tree.heading('status', text='Status')
        self.tree.column('status', width=100)
        
        self.tree.heading('start', text='Início')
        self.tree.column('start', width=150)
        
        self.tree.heading('end', text='Fim')
        self.tree.column('end', width=150)
        
        self.tree.heading('duration', text='Duração Total')
        self.tree.column('duration', width=100)
        
        # Scrollbars
        y_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        x_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky=tk.NSEW)
        y_scrollbar.grid(row=0, column=1, sticky=tk.NS)
        x_scrollbar.grid(row=1, column=0, sticky=tk.EW)
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Status bar
        self.status_label = ttk.Label(main_frame, text="", anchor=tk.W)
        self.status_label.pack(fill=tk.X, pady=(10, 0))
        
    def _load_logs(self):
        """Carrega os logs do task manager"""
        self._apply_filters()
        
    def _apply_filters(self):
        """Aplica os filtros aos logs"""
        # Limpar treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Obter filtros
        text_filter = self.text_filter.get().lower()
        date_from_str = self.date_from.get()
        date_to_str = self.date_to.get()
        
        try:
            date_from = datetime.strptime(date_from_str, "%Y-%m-%d") if date_from_str else None
            date_to = datetime.strptime(date_to_str, "%Y-%m-%d") if date_to_str else None
            if date_to:
                date_to = date_to.replace(hour=23, minute=59, second=59)
        except ValueError:
            messagebox.showerror("Erro", "Formato de data inválido. Use YYYY-MM-DD")
            return
        
        # Filtrar tarefas
        filtered_tasks = []
        all_tasks = self.task_manager.get_all_tasks()
        
        for task in all_tasks:
            # Filtro por texto
            if text_filter:
                search_text = f"{task.ticket} {task.email} {task.category} {task.description}".lower()
                if text_filter not in search_text:
                    continue
            
            # Filtro por data
            if date_from or date_to:
                task_date = task.finished_at or task.created_at
                if date_from and task_date < date_from:
                    continue
                if date_to and task_date > date_to:
                    continue
            
            filtered_tasks.append(task)
        
        # Popular treeview
        for task in filtered_tasks:
            priority_color = ThemeManager.get_priority_color(task.priority)
            
            start_str = task.started_at.strftime("%d/%m/%Y %H:%M") if task.started_at else "-"
            end_str = task.finished_at.strftime("%d/%m/%Y %H:%M") if task.finished_at else "-"
            
            values = (
                task.ticket,
                task.email,
                task.category,
                ThemeManager.get_priority_name(task.priority),
                task.status,
                start_str,
                end_str,
                task.get_total_time_str()
            )
            
            item_id = self.tree.insert('', tk.END, values=values)
            # Aplicar cor na célula de prioridade
            self.tree.item(item_id, tags=(task.priority,))
        
        # Configurar tags de cores
        self.tree.tag_configure('baixa', foreground=ThemeManager.COLORS['priority_baixa'])
        self.tree.tag_configure('media', foreground=ThemeManager.COLORS['priority_media'])
        self.tree.tag_configure('alta', foreground=ThemeManager.COLORS['priority_alta'])
        self.tree.tag_configure('urgente', foreground=ThemeManager.COLORS['priority_urgente'])
        
        # Atualizar status
        self.status_label.config(text=f"Total: {len(filtered_tasks)} chamado(s) encontrado(s)")
        
    def _clear_filters(self):
        """Limpa todos os filtros"""
        self.text_filter.delete(0, tk.END)
        self.date_from.delete(0, tk.END)
        self.date_from.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.date_to.delete(0, tk.END)
        self.date_to.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self._apply_filters()
        
    def _export_to_excel(self):
        """Exporta os logs filtrados para Excel"""
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            messagebox.showwarning("Dependência Ausente", 
                                 "Para exportar para Excel, instale: pip install openpyxl\n\n"
                                 "Os dados serão exportados em formato CSV.")
            self._export_to_csv()
            return
        
        # Gerar nome do arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Gestao_Atividades_{timestamp}.xlsx"
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Chamados"
        
        # Cabeçalhos
        headers = ['Ticket', 'E-mail', 'Categoria', 'Descrição', 'Prioridade', 
                  'Status', 'Criado Em', 'Iniciado Em', 'Finalizado Em', 'Duração Total']
        ws.append(headers)
        
        # Dados
        all_tasks = self.task_manager.get_all_tasks()
        for task in all_tasks:
            ws.append([
                task.ticket,
                task.email,
                task.category,
                task.description,
                ThemeManager.get_priority_name(task.priority),
                task.status,
                task.created_at.strftime("%d/%m/%Y %H:%M:%S"),
                task.started_at.strftime("%d/%m/%Y %H:%M:%S") if task.started_at else "",
                task.finished_at.strftime("%d/%m/%Y %H:%M:%S") if task.finished_at else "",
                task.get_total_time_str()
            ])
        
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="0078d4", end_color="0078d4", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        # Ajustar largura das colunas
        column_widths = [15, 30, 35, 50, 12, 15, 20, 20, 20, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Salvar arquivo
        save_path = CONFIG_DIR / filename
        wb.save(str(save_path))
        
        messagebox.showinfo("Sucesso", f"Arquivo exportado com sucesso!\n\nLocal: {save_path}")
        
    def _export_to_csv(self):
        """Exporta os logs para CSV como fallback"""
        import csv
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Gestao_Atividades_{timestamp}.csv"
        save_path = CONFIG_DIR / filename
        
        all_tasks = self.task_manager.get_all_tasks()
        
        with open(save_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Ticket', 'E-mail', 'Categoria', 'Descrição', 'Prioridade', 
                           'Status', 'Criado Em', 'Iniciado Em', 'Finalizado Em', 'Duração Total'])
            
            for task in all_tasks:
                writer.writerow([
                    task.ticket,
                    task.email,
                    task.category,
                    task.description,
                    ThemeManager.get_priority_name(task.priority),
                    task.status,
                    task.created_at.strftime("%d/%m/%Y %H:%M:%S"),
                    task.started_at.strftime("%d/%m/%Y %H:%M:%S") if task.started_at else "",
                    task.finished_at.strftime("%d/%m/%Y %H:%M:%S") if task.finished_at else "",
                    task.get_total_time_str()
                ])
        
        messagebox.showinfo("Sucesso", f"Arquivo CSV exportado com sucesso!\n\nLocal: {save_path}")


class TaskManager:
    """Gerencia todas as tarefas do sistema"""
    
    def __init__(self):
        self.tasks = {}
        self.notification_manager = NotificationManager()
        self._load_tasks()
        
    def _load_tasks(self):
        """Carrega tarefas salvas"""
        if TASKS_FILE.exists():
            try:
                with open(TASKS_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    for task_data in data:
                        task = Task.from_dict(task_data)
                        self.tasks[task.ticket] = task
            except Exception as e:
                print(f"Erro ao carregar tarefas: {e}")
                
    def _save_tasks(self):
        """Salva tarefas em arquivo"""
        try:
            with open(TASKS_FILE, 'w', encoding='utf-8') as f:
                json.dump([task.to_dict() for task in self.tasks.values()], f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Erro ao salvar tarefas: {e}")
            
    def add_task(self, ticket, email, category, description, priority="media"):
        """Adiciona uma nova tarefa"""
        if ticket in self.tasks:
            return False, "Ticket já existe!"
            
        task = Task(ticket, email, category, description, priority)
        self.tasks[ticket] = task
        self._save_tasks()
        return True, "Tarefa adicionada com sucesso!"
        
    def start_task(self, ticket):
        """Inicia uma tarefa"""
        if ticket not in self.tasks:
            return False, "Tarefa não encontrada!"
            
        task = self.tasks[ticket]
        task.start()
        self._save_tasks()
        return True, f"Tarefa {ticket} iniciada!"
        
    def pause_task(self, ticket):
        """Pausa uma tarefa"""
        if ticket not in self.tasks:
            return False, "Tarefa não encontrada!"
            
        task = self.tasks[ticket]
        task.pause()
        self._save_tasks()
        return True, f"Tarefa {ticket} pausada!"
        
    def resume_task(self, ticket):
        """Retoma uma tarefa"""
        if ticket not in self.tasks:
            return False, "Tarefa não encontrada!"
            
        task = self.tasks[ticket]
        task.resume()
        self._save_tasks()
        return True, f"Tarefa {ticket} retomada!"
        
    def finish_task(self, ticket):
        """Conclui uma tarefa"""
        if ticket not in self.tasks:
            return False, "Tarefa não encontrada!"
            
        task = self.tasks[ticket]
        task.finish()
        self._save_tasks()
        return True, f"Tarefa {ticket} concluída!"
        
    def delete_task(self, ticket):
        """Remove uma tarefa"""
        if ticket not in self.tasks:
            return False, "Tarefa não encontrada!"
            
        del self.tasks[ticket]
        self._save_tasks()
        return True, "Tarefa removida!"
        
    def get_active_tasks(self):
        """Retorna todas as tarefas ativas (em andamento ou pausadas)"""
        return [t for t in self.tasks.values() if t.status in ['em_andamento', 'pausado']]
        
    def get_all_tasks(self):
        """Retorna todas as tarefas"""
        return list(self.tasks.values())
        
    def update_priorities_in_background(self):
        """Thread para verificar múltiplas tarefas e notificar"""
        while True:
            active_tasks = self.get_active_tasks()
            self.notification_manager.check_multiple_tasks(active_tasks)
            time.sleep(60)  # Verifica a cada minuto


class SystemTrayIcon:
    """Gerencia o ícone na bandeja do sistema"""
    
    def __init__(self, root, task_manager):
        self.root = root
        self.task_manager = task_manager
        self.tray_icon = None
        self.menu = None
        self._create_tray_icon()
        
    def _create_tray_icon(self):
        """Cria o ícone na bandeja"""
        # Criar menu de contexto
        self.menu = tk.Menu(self.root, tearoff=0)
        self.menu.add_command(label="Abrir Gestão de Atividades", command=self._show_window)
        self.menu.add_command(label="Minimizar", command=self._minimize_window)
        self.menu.add_separator()
        self.menu.add_command(label="Sair", command=self._exit_app)
        
        # Nota: Implementação completa do system tray requer bibliotecas adicionais
        # Para simplicidade, usamos abordagem com protocolo WM_DELETE_WINDOW
        
    def _show_window(self):
        """Mostra a janela principal"""
        self.root.deiconify()
        self.root.lift()
        self.root.focus_force()
        
    def _minimize_window(self):
        """Minimiza a janela"""
        self.root.withdraw()
        
    def _exit_app(self):
        """Sai da aplicação"""
        self.root.quit()


class GestaoDeAtividadesApp:
    """Aplicação principal"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("📋 Gestão de Atividades")
        self.root.geometry("1100x700")
        self.root.minsize(900, 600)
        
        # Configurar ícone (se existir)
        try:
            self.root.iconbitmap("icon.ico")
        except:
            pass
        
        # Tema escuro
        self.root.configure(bg=ThemeManager.COLORS['bg_dark'])
        
        # Gerenciador de tarefas
        self.task_manager = TaskManager()
        
        # System tray
        self.tray = SystemTrayIcon(self.root, self.task_manager)
        
        # Configurar estilos
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self._configure_styles()
        
        # Variáveis
        self.selected_task = None
        
        # Criar interface
        self._create_menu()
        self._create_widgets()
        
        # Iniciar thread de notificações
        self.notification_thread = threading.Thread(
            target=self.task_manager.update_priorities_in_background,
            daemon=True
        )
        self.notification_thread.start()
        
        # Configurar protocolo de fechamento
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        
        # Carregar dados salvos
        self._refresh_task_list()
        
    def _configure_styles(self):
        """Configura os estilos dos widgets"""
        colors = ThemeManager.COLORS
        
        self.style.configure("TFrame", background=colors['bg_dark'])
        self.style.configure("TLabel", 
                            background=colors['bg_dark'], 
                            foreground=colors['fg_primary'],
                            font=('Segoe UI', 10))
        self.style.configure("TButton",
                            background=colors['accent'],
                            foreground=colors['fg_primary'],
                            font=('Segoe UI', 10, 'bold'),
                            padding=5)
        self.style.map("TButton",
                      background=[('active', colors['accent_hover'])])
        self.style.configure("TEntry",
                            fieldbackground=colors['bg_medium'],
                            foreground=colors['fg_primary'],
                            font=('Segoe UI', 10),
                            padding=5)
        self.style.configure("TCombobox",
                            fieldbackground=colors['bg_medium'],
                            foreground=colors['fg_primary'],
                            font=('Segoe UI', 10))
        self.style.configure("TLabelframe",
                            background=colors['bg_dark'],
                            foreground=colors['fg_primary'])
        self.style.configure("TLabelframe.Label",
                            background=colors['bg_dark'],
                            foreground=colors['fg_primary'],
                            font=('Segoe UI', 11, 'bold'))
                            
    def _create_menu(self):
        """Cria a barra de menus"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # Menu Arquivo
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Arquivo", menu=file_menu)
        file_menu.add_command(label="Ver Histórico de Logs", command=self._open_logs_window)
        file_menu.add_separator()
        file_menu.add_command(label="Sair", command=self._on_close)
        
        # Menu Ajuda
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Ajuda", menu=help_menu)
        help_menu.add_command(label="Sobre", command=self._show_about)
        
    def _create_widgets(self):
        """Cria os widgets principais"""
        colors = ThemeManager.COLORS
        
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Frame superior - Adicionar tarefa
        add_frame = ttk.LabelFrame(main_frame, text="➕ Novo Chamado", padding="10")
        add_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Linha 1 - Ticket e E-mail
        line1 = ttk.Frame(add_frame)
        line1.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(line1, text="Ticket:").pack(side=tk.LEFT, padx=(0, 5))
        self.ticket_entry = ttk.Entry(line1, width=20)
        self.ticket_entry.pack(side=tk.LEFT, padx=(0, 15))
        self.ticket_entry.bind('<FocusOut>', lambda e: self._auto_fill_from_clipboard())
        
        ttk.Label(line1, text="E-mail:").pack(side=tk.LEFT, padx=(0, 5))
        self.email_entry = ttk.Entry(line1, width=35)
        self.email_entry.pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Label(line1, text="Prioridade:").pack(side=tk.LEFT, padx=(0, 5))
        self.priority_combo = ttk.Combobox(line1, values=["baixa", "media", "alta", "urgente"], 
                                          state="readonly", width=10)
        self.priority_combo.set("media")
        self.priority_combo.pack(side=tk.LEFT)
        
        # Linha 2 - Categoria
        line2 = ttk.Frame(add_frame)
        line2.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(line2, text="Categoria:").pack(side=tk.LEFT, padx=(0, 5))
        self.category_entry = ttk.Entry(line2, width=70)
        self.category_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        # Linha 3 - Descrição
        line3 = ttk.Frame(add_frame)
        line3.pack(fill=tk.X)
        
        ttk.Label(line3, text="Descrição:").pack(side=tk.LEFT, padx=(0, 5), anchor=tk.NW)
        self.desc_text = tk.Text(line3, height=3, width=70, 
                                bg=colors['bg_medium'], 
                                fg=colors['fg_primary'],
                                font=('Segoe UI', 10),
                                insertbackground=colors['fg_primary'])
        self.desc_text.pack(side=tk.LEFT, padx=(0, 5))
        
        # Botão adicionar
        btn_add = ttk.Button(add_frame, text="✅ Adicionar Tarefa", command=self._add_task)
        btn_add.pack(pady=(10, 0))
        
        # Frame do meio - Lista de tarefas
        list_frame = ttk.LabelFrame(main_frame, text="📝 Tarefas em Andamento", padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Treeview
        columns = ('ticket', 'email', 'priority', 'status', 'time', 'actions')
        self.task_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=10)
        
        self.task_tree.heading('ticket', text='Ticket')
        self.task_tree.column('ticket', width=120)
        
        self.task_tree.heading('email', text='E-mail')
        self.task_tree.column('email', width=200)
        
        self.task_tree.heading('priority', text='Prioridade')
        self.task_tree.column('priority', width=80)
        
        self.task_tree.heading('status', text='Status')
        self.task_tree.column('status', width=100)
        
        self.task_tree.heading('time', text='Tempo')
        self.task_tree.column('time', width=100)
        
        self.task_tree.heading('actions', text='Ações')
        self.task_tree.column('actions', width=200)
        
        # Scrollbar
        y_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.task_tree.yview)
        self.task_tree.configure(yscrollcommand=y_scrollbar.set)
        
        self.task_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind de seleção
        self.task_tree.bind('<<TreeviewSelect>>', self._on_task_select)
        
        # Frame inferior - Ações
        actions_frame = ttk.LabelFrame(main_frame, text="⚙️ Ações", padding="10")
        actions_frame.pack(fill=tk.X)
        
        self.btn_start = ttk.Button(actions_frame, text="▶️ Iniciar", command=self._start_task, state=tk.DISABLED)
        self.btn_start.pack(side=tk.LEFT, padx=(0, 5))
        
        self.btn_pause = ttk.Button(actions_frame, text="⏸️ Pausar", command=self._pause_task, state=tk.DISABLED)
        self.btn_pause.pack(side=tk.LEFT, padx=(0, 5))
        
        self.btn_resume = ttk.Button(actions_frame, text="⏯️ Retomar", command=self._resume_task, state=tk.DISABLED)
        self.btn_resume.pack(side=tk.LEFT, padx=(0, 5))
        
        self.btn_finish = ttk.Button(actions_frame, text="✅ Concluir", command=self._finish_task, state=tk.DISABLED)
        self.btn_finish.pack(side=tk.LEFT, padx=(0, 5))
        
        self.btn_delete = ttk.Button(actions_frame, text="🗑️ Excluir", command=self._delete_task, state=tk.DISABLED)
        self.btn_delete.pack(side=tk.LEFT, padx=(0, 5))
        
        # Label de status
        self.status_label = ttk.Label(actions_frame, text="", anchor=tk.W)
        self.status_label.pack(side=tk.RIGHT)
        
        # Atualizar tempos periodicamente
        self._update_times()
        
    def _auto_fill_from_clipboard(self):
        """Tenta preencher automaticamente do clipboard quando o ticket perde o foco"""
        try:
            ticket = self.ticket_entry.get().strip()
            if ticket and not hasattr(self, '_filled_from_clipboard'):
                # Verifica se parece um ticket válido
                if ticket.startswith('#TI') or ticket.startswith('TI'):
                    self._filled_from_clipboard = True
        except:
            pass
            
    def _parse_call_line(self, line):
        """Analisa a linha do chamado e extrai informações"""
        parts = line.split('\t')
        
        if len(parts) >= 4:
            ticket = parts[0].strip()
            email = parts[1].strip()
            category = parts[2].strip()
            description = parts[3].strip().strip('"')
            
            return ticket, email, category, description
        
        return None, None, None, None
        
    def _add_task(self):
        """Adiciona uma nova tarefa"""
        ticket = self.ticket_entry.get().strip()
        email = self.email_entry.get().strip()
        category = self.category_entry.get().strip()
        description = self.desc_text.get("1.0", tk.END).strip()
        priority = self.priority_combo.get()
        
        if not ticket:
            messagebox.showerror("Erro", "Informe o número do ticket!")
            return
            
        if not email:
            messagebox.showerror("Erro", "Informe o e-mail do solicitante!")
            return
            
        success, message = self.task_manager.add_task(ticket, email, category, description, priority)
        
        if success:
            messagebox.showinfo("Sucesso", message)
            self._clear_form()
            self._refresh_task_list()
        else:
            messagebox.showerror("Erro", message)
            
    def _clear_form(self):
        """Limpa o formulário"""
        self.ticket_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
        self.category_entry.delete(0, tk.END)
        self.desc_text.delete("1.0", tk.END)
        self.priority_combo.set("media")
        self._filled_from_clipboard = False
        
    def _refresh_task_list(self):
        """Atualiza a lista de tarefas"""
        # Limpar treeview
        for item in self.task_tree.get_children():
            self.task_tree.delete(item)
        
        # Adicionar tarefas
        tasks = self.task_manager.get_all_tasks()
        
        for task in tasks:
            priority_color = ThemeManager.get_priority_color(task.priority)
            status_names = {
                'pendente': '⏳ Pendente',
                'em_andamento': '▶️ Em Andamento',
                'pausado': '⏸️ Pausado',
                'concluido': '✅ Concluído'
            }
            
            values = (
                task.ticket,
                task.email,
                ThemeManager.get_priority_name(task.priority),
                status_names.get(task.status, task.status),
                task.get_total_time_str(),
                ""
            )
            
            item_id = self.task_tree.insert('', tk.END, values=values, tags=(task.priority,))
            
            # Armazenar referência à tarefa no item
            self.task_tree.item(item_id, values=values)
        
        # Configurar tags de cores
        self.task_tree.tag_configure('baixa', foreground=ThemeManager.COLORS['priority_baixa'])
        self.task_tree.tag_configure('media', foreground=ThemeManager.COLORS['priority_media'])
        self.task_tree.tag_configure('alta', foreground=ThemeManager.COLORS['priority_alta'])
        self.task_tree.tag_configure('urgente', foreground=ThemeManager.COLORS['priority_urgente'])
        
    def _on_task_select(self, event):
        """Handle seleção de tarefa"""
        selection = self.task_tree.selection()
        
        if selection:
            item = selection[0]
            values = self.task_tree.item(item, 'values')
            ticket = values[0]
            
            # Encontrar a tarefa
            task = self.task_manager.tasks.get(ticket)
            
            if task:
                self.selected_task = ticket
                self._enable_action_buttons(task.status)
            else:
                self.selected_task = None
                self._disable_action_buttons()
        else:
            self.selected_task = None
            self._disable_action_buttons()
            
    def _enable_action_buttons(self, status):
        """Habilita os botões de ação baseados no status"""
        self.btn_start.config(state=tk.NORMAL if status == 'pendente' else tk.DISABLED)
        self.btn_pause.config(state=tk.NORMAL if status == 'em_andamento' else tk.DISABLED)
        self.btn_resume.config(state=tk.NORMAL if status == 'pausado' else tk.DISABLED)
        self.btn_finish.config(state=tk.NORMAL if status in ['em_andamento', 'pausado'] else tk.DISABLED)
        self.btn_delete.config(state=tk.NORMAL)
        
    def _disable_action_buttons(self):
        """Desabilita todos os botões de ação"""
        self.btn_start.config(state=tk.DISABLED)
        self.btn_pause.config(state=tk.DISABLED)
        self.btn_resume.config(state=tk.DISABLED)
        self.btn_finish.config(state=tk.DISABLED)
        self.btn_delete.config(state=tk.DISABLED)
        
    def _start_task(self):
        """Inicia a tarefa selecionada"""
        if not self.selected_task:
            return
            
        success, message = self.task_manager.start_task(self.selected_task)
        
        if success:
            self.status_label.config(text=message)
            self._refresh_task_list()
            self._on_task_select(None)  # Atualizar botões
        else:
            messagebox.showerror("Erro", message)
            
    def _pause_task(self):
        """Pausa a tarefa selecionada"""
        if not self.selected_task:
            return
            
        success, message = self.task_manager.pause_task(self.selected_task)
        
        if success:
            self.status_label.config(text=message)
            self._refresh_task_list()
            self._on_task_select(None)
        else:
            messagebox.showerror("Erro", message)
            
    def _resume_task(self):
        """Retoma a tarefa selecionada"""
        if not self.selected_task:
            return
            
        success, message = self.task_manager.resume_task(self.selected_task)
        
        if success:
            self.status_label.config(text=message)
            self._refresh_task_list()
            self._on_task_select(None)
        else:
            messagebox.showerror("Erro", message)
            
    def _finish_task(self):
        """Conclui a tarefa selecionada"""
        if not self.selected_task:
            return
            
        if messagebox.askyesno("Confirmar", f"Deseja realmente concluir o chamado {self.selected_task}?"):
            success, message = self.task_manager.finish_task(self.selected_task)
            
            if success:
                self.status_label.config(text=message)
                self._refresh_task_list()
                self._on_task_select(None)
            else:
                messagebox.showerror("Erro", message)
                
    def _delete_task(self):
        """Exclui a tarefa selecionada"""
        if not self.selected_task:
            return
            
        if messagebox.askyesno("Confirmar", f"Deseja realmente excluir o chamado {self.selected_task}?\n\nEsta ação não pode ser desfeita!"):
            success, message = self.task_manager.delete_task(self.selected_task)
            
            if success:
                self.status_label.config(text=message)
                self.selected_task = None
                self._refresh_task_list()
                self._disable_action_buttons()
            else:
                messagebox.showerror("Erro", message)
                
    def _update_times(self):
        """Atualiza os tempos das tarefas periodicamente"""
        self._refresh_task_list()
        self.root.after(1000, self._update_times)  # Atualiza a cada segundo
        
    def _open_logs_window(self):
        """Abre a janela de logs"""
        logs_window = LogsWindow(self.root, self.task_manager)
        logs_window.transient(self.root)
        logs_window.grab_set()
        
    def _show_about(self):
        """Mostra diálogo sobre"""
        messagebox.showinfo("Sobre - Gestão de Atividades",
                          "Gestão de Atividades v1.0.0\n\n"
                          "Sistema de gerenciamento de tarefas para TI\n\n"
                          "Funcionalidades:\n"
                          "• Controle de múltiplas tarefas em paralelo\n"
                          "• Notificações de 30 em 30 minutos\n"
                          "• Pause/Resume com cálculo de tempo\n"
                          "• Histórico completo com filtros\n"
                          "• Exportação para Excel\n\n"
                          "Desenvolvido por: Isaacssjr")
                          
    def _on_close(self):
        """Handle fechamento da aplicação"""
        if messagebox.askyesno("Sair", "Deseja realmente sair do Gestão de Atividades?\n\nO sistema continuará rodando na bandeja do sistema."):
            self.root.quit()
            self.root.destroy()
        else:
            self.root.withdraw()  # Minimiza para tray
            
    def run(self):
        """Inicia a aplicação"""
        self.root.mainloop()


def main():
    """Função principal"""
    app = GestaoDeAtividadesApp()
    app.run()


if __name__ == "__main__":
    main()
